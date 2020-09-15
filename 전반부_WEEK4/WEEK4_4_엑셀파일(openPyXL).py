#OpenPyXL 연습하기

import openpyxl

wb = openpyxl.Workbook()      #새로운 워크북(엑셀)을 만들기
sheet = wb.active             #워크북.active   ->  현재 활성화된 엑셀 파일의 시트 선택
sheet['A1'] = "Hello World"   #시트[셀번호]  ->  셀번호를 통해 셀 값 입력, 변경   (셀번호를 알아야하므로 여러개 데이터 입력할때는 불편함)
sheet.cell(row=3, column=3).value = "Good Bye"   #시트.cell(열, 행).value  ->  원하는 데이터를 원하는 위치에 넣기 편리

sheet.append(["Python", "Java", "HTML", "JAVASCRIPT"])   #시트.append  ->  리스트 이용하여 데이터 위치와는 상관없이 가장 아래에 데이터 저장. 행별로 추가.
sheet.append(["Coala", "Study", "Crawling"])
wb.save("test2.xlsx")     #해당하는 워크북을 원하는 이름으로 저장
