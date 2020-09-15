##복습문제 박지연
'''
#1. csv
import requests
from bs4 import BeautifulSoup

f = open("navermusic.csv", "w")
f.write("순위, 곡명, 아티스트\n")

for i in range(1, 3):
    raw = requests.get("https://music.naver.com/chart/musicianLeague.nhn?duration=DAILY&page="+str(i), headers = {"User-Agent":"Mozilla/5.0"})
    html = BeautifulSoup(raw.text, 'html.parser')

    music = html.select("tr._tracklist_move")

    for mus in music:
        ranking = mus.select_one("span.num").text.strip()
        title = mus.select_one("td.tb_name span.tit").text.strip()
        artist = mus.select_one("td.tb_artist span.tit").text.strip()
        print(ranking, "/", title, "/", artist)

        title = title.replace(",", "")
        artist = artist.replace(",", "")

        f.write(ranking + ',' + title + ',' + artist + '\n')

f.close()

'''
'''
#2. excel
import requests
from bs4 import BeautifulSoup

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(["순위", "곡명", "아티스트"])  

for i in range(1, 3):
    raw = requests.get("https://music.naver.com/chart/musicianLeague.nhn?duration=DAILY&page="+str(i), headers = {"User-Agent":"Mozilla/5.0"})
    html = BeautifulSoup(raw.text, 'html.parser')

    music = html.select("tr._tracklist_move")

    for mus in music:
        ranking = mus.select_one("span.num").text.strip()
        title = mus.select_one("td.tb_name span.tit").text.strip()
        artist = mus.select_one("td.tb_artist span.tit").text.strip()
        print(ranking, "/", title, "/", artist)

        sheet.append([ranking, title, artist])
wb.save("navermusic.xlsx")
'''
'''
##성혜원

import requests
from bs4 import BeautifulSoup
import openpyxl

days = input("mon / tue / wed / thu / fri / sat / sun 중 요일을 고르시오: ")

try:
    wb = openpyxl.load_workbook("naverwebtoon.xlsx")
    sheet = wb.active
    print("불러오기 완료")
except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["요일", "제목", "작가"])
    print("새로운 파일을 만들었습니다.")

for day in days:
    raw = requests.get("https://comic.naver.com/webtoon/weekdayList.nhn?week="+days, headers = {"User-Agent":"Mozilla/5.0"})
    html = BeautifulSoup(raw.text, 'html.parser')

    webtoon = html.select("div.list_area li")

    for toon in webtoon:
        title = toon.select_one("ul.img_list dl>dt>a").text.strip()
        writer = toon.select_one("dd.desc > a").text.strip()
        print(title, "/", writer)

        sheet.append([days, title, writer])

wb.save("naverwebtoon.xlsx")
'''
'''
##복습문제 윤희준
import requests
from bs4 import BeautifulSoup

f = open("caltech.csv", "w", encoding="UTF-8")
f.write("기사 날짜, 제목\n")

for i in range(1, 11):
    raw = requests.get("https://www.caltech.edu/about/news?p="+str(i), headers = {"User-Agent":"Mozilla/5.0"})
    html = BeautifulSoup(raw.text, 'html.parser')

    article = html.select("div.article-teaser__info")

    for artcl in article:
        date = artcl.select_one("span.article-teaser__published-date__date").text.strip()
        title = artcl.select_one("div.article-teaser__title").text.strip()
        print(date, "/", title)
    
        date = date.replace(",", "")
        title = title.replace(",", "")

        f.write(date + ',' + title + '\n')

f.close()
'''

##복습 문제 이미경

import requests
from bs4 import BeautifulSoup

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(["내용"])

raw = requests.get("https://www.naver.com/", headers={"User-Agent": "Mozilla/5.0"})
html = BeautifulSoup(raw.text, 'html.parser')

items = html.select("ul.list_nav.NM_FAVORITE_LIST>li")

for it in items :
    item = it.select_one("a.nav").text.strip()
    print(item)
    sheet.append([item])

#  ul.NM_FAVORITE_LIST li.nav_item a

wb.save("naveritem.xlsx")

